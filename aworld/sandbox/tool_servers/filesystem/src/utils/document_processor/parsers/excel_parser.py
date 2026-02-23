"""Excel 解析器"""

import logging
from pathlib import Path
from typing import Any, Dict, Optional

try:
    import pandas as pd
    from openpyxl import load_workbook
except ImportError:
    pd = None
    load_workbook = None

from ..base_parser import BaseParser

logger = logging.getLogger(__name__)


class ExcelParser(BaseParser):
    def get_supported_types(self) -> list[str]:
        return ["xlsx", "xls"]

    def can_handle(self, file_type: str) -> bool:
        return file_type.lower() in ["xlsx", "xls"]

    async def parse(
        self,
        file_path: Path,
        task_id: str = "",
        source_file_id: Optional[str] = None,
        source_file_name: Optional[str] = None,
        afts_service: Any = None,
        output_path: Optional[Path] = None,
        **kwargs,
    ) -> Dict[str, Any]:
        if pd is None or load_workbook is None:
            raise RuntimeError("未安装 pandas 或 openpyxl。请安装: pip install pandas openpyxl")
        markdown_content = await self._extract_xlsx_content(file_path)
        source_name = source_file_name or file_path.stem
        parsed_file_path = await self._save_markdown_to_file(
            markdown_content, task_id, source_name, output_path=output_path
        )
        return {"file_path": parsed_file_path}

    async def _extract_xlsx_content(self, file_path: Path) -> str:
        workbook = load_workbook(file_path, data_only=True)
        parts = [f"# {file_path.stem}\n\n", f"**工作表数**: {len(workbook.sheetnames)}\n\n", "---\n\n"]
        for sheet_name in workbook.sheetnames:
            parts.append(f"## 工作表: {sheet_name}\n\n")
            try:
                ws = workbook[sheet_name]
                if ws.max_row == 0 or ws.max_column == 0:
                    parts.append("*工作表为空*\n\n")
                    continue
                merged = await self._process_merged_cells(ws)
                data = []
                for row_idx, row in enumerate(
                    ws.iter_rows(
                        min_row=1,
                        max_row=ws.max_row,
                        min_col=1,
                        max_col=ws.max_column,
                        values_only=True,
                    )
                ):
                    row_data = []
                    for col_idx, val in enumerate(row):
                        if (row_idx, col_idx) in merged:
                            val = "" if merged[(row_idx, col_idx)] is None else merged[(row_idx, col_idx)]
                        if val is None or (isinstance(val, float) and pd.isna(val)):
                            val = ""
                        else:
                            val = str(val).strip()
                        row_data.append(val)
                    data.append(row_data)
                if not data:
                    parts.append("*工作表为空*\n\n")
                    continue
                df = (
                    pd.DataFrame(data)
                    .replace("", None)
                    .dropna(how="all")
                    .dropna(axis=1, how="all")
                    .fillna("")
                )
                if df.empty:
                    parts.append("*工作表为空*\n\n")
                else:
                    headers = [str(c) for c in df.columns]
                    parts.append("| " + " | ".join(headers) + " |\n")
                    parts.append("|" + "---|" * len(headers) + "\n")
                    for _, row in df.iterrows():
                        parts.append("| " + " | ".join(str(v) if v != "" else "" for v in row) + " |\n")
                    parts.append("\n")
            except Exception as e:
                logger.warning("excel_parser sheet %s error: %s", sheet_name, e)
                parts.append(f"**错误**: {e}\n\n")
        workbook.close()
        return "".join(parts)

    async def _process_merged_cells(self, worksheet) -> dict:
        out = {}
        for merged_range in worksheet.merged_cells.ranges:
            min_col, min_row, max_col_idx, max_row_idx = merged_range.bounds
            min_col -= 1
            min_row -= 1
            max_col_idx -= 1
            max_row_idx -= 1
            main_val = worksheet.cell(min_row + 1, min_col + 1).value or ""
            for row in range(min_row, max_row_idx + 1):
                for col in range(min_col, max_col_idx + 1):
                    out[(row, col)] = None if (row != min_row or col != min_col) else main_val
        return out

